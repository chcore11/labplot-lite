from pathlib import Path
import re
import uuid

from flask import Flask, render_template, request, send_file
from openpyxl import Workbook

import pandas as pd
import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
SAMPLE_DIR = BASE_DIR / "samples"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
SAMPLE_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".csv", ".xls", ".xlsx"}

SAMPLE_FILES = {
    "sample_01_temp_c_to_k.xlsx": "示例 1：摄氏度转开尔文",
    "sample_02_voltage_scale.xlsx": "示例 2：电压比例换算",
    "sample_03_abs_error.xlsx": "示例 3：绝对误差",
    "sample_04_invalid_log.xlsx": "示例 4：非法 log 拦截",
}

CHART_TYPES = {
    "line": "折线图",
    "scatter": "散点图",
    "line_marker": "散点 + 连线",
}

CALC_TEMPLATES = {
    "multiply": "两列相乘 A × B",
    "divide": "两列相除 A ÷ B",
    "add": "两列相加 A + B",
    "subtract": "两列相减 A - B",
    "square": "某列平方 A²",
    "add_const": "某列加常数 A + k",
    "subtract_const": "某列减常数 A - k",
    "multiply_const": "某列乘常数 A × k",
    "divide_const": "某列除常数 A ÷ k",
    "sqrt": "平方根 sqrt(A)",
    "log10": "常用对数 log10(A)",
    "ln": "自然对数 ln(A)",
    "abs": "绝对值 abs(A)",
}

FIT_TYPES = {
    "none": "不拟合",
    "linear": "一次线性拟合 y = ax + b",
    "quadratic": "二次拟合 y = ax² + bx + c",
}

METRIC_MODES = {
    "basic": "基础指标（推荐）",
    "custom": "自定义选择",
}

AVAILABLE_METRICS = {
    "r2": "R²",
    "rmse": "RMSE",
    "mae": "MAE",
    "max_abs_error": "最大绝对误差",
    "mse": "MSE",
    "residual_mean": "残差均值",
    "residual_std": "残差标准差",
}

BASIC_METRICS = ["r2", "rmse", "mae", "max_abs_error"]

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024


# =========================
# Matplotlib 中文字体配置
# =========================
def setup_matplotlib_chinese_font():
    preferred_fonts = [
        "Microsoft YaHei",
        "SimHei",
        "SimSun",
        "PingFang SC",
        "Heiti SC",
        "Noto Sans CJK SC",
        "WenQuanYi Zen Hei",
        "Arial Unicode MS",
    ]

    available_fonts = {f.name for f in font_manager.fontManager.ttflist}

    for font_name in preferred_fonts:
        if font_name in available_fonts:
            plt.rcParams["font.sans-serif"] = [font_name]
            plt.rcParams["axes.unicode_minus"] = False
            return font_name

    plt.rcParams["axes.unicode_minus"] = False
    return None


CHINESE_FONT_NAME = setup_matplotlib_chinese_font()
print("当前 Matplotlib 中文字体：", CHINESE_FONT_NAME)


# =========================
# 示例文件与临时文件管理
# =========================
def write_sample_xlsx(filename: str, headers: list[str], rows: list[list]):
    path = SAMPLE_DIR / filename

    if path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "sample"

    ws["A1"] = "LabPlot Lite 示例数据"
    ws["A2"] = "说明：真实表头在第 9 行"
    ws["A3"] = "说明：真实数据从第 10 行开始"
    ws["A4"] = "你可以用这个文件测试上传、表头识别、计算列和绘图功能"

    header_row = 9

    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_index, value=header)

    for row_index, row_data in enumerate(rows, start=10):
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    wb.save(path)


def ensure_sample_files():
    write_sample_xlsx(
        "sample_01_temp_c_to_k.xlsx",
        ["temp_c", "sensor_value"],
        [
            [-20, 1.2],
            [-10, 1.5],
            [0, 1.8],
            [10, 2.1],
            [20, 2.5],
            [30, 2.9],
            [40, 3.3],
            [50, 3.8],
        ],
    )

    write_sample_xlsx(
        "sample_02_voltage_scale.xlsx",
        ["voltage_v", "current_a"],
        [
            [0.5, 0.02],
            [1.0, 0.04],
            [1.5, 0.06],
            [2.0, 0.08],
            [2.5, 0.10],
            [3.0, 0.12],
            [3.3, 0.13],
            [5.0, 0.20],
        ],
    )

    write_sample_xlsx(
        "sample_03_abs_error.xlsx",
        ["sample_id", "error"],
        [
            [1, -0.12],
            [2, 0.08],
            [3, -0.31],
            [4, 0.00],
            [5, 0.15],
            [6, -0.49],
            [7, 0.21],
            [8, -0.05],
        ],
    )

    write_sample_xlsx(
        "sample_04_invalid_log.xlsx",
        ["sample_id", "raw_value"],
        [
            [1, 10],
            [2, 5],
            [3, 1],
            [4, 0],
            [5, -1],
            [6, 20],
            [7, -5],
            [8, 100],
        ],
    )


def cleanup_temp_files():
    deleted_count = 0

    for folder in [UPLOAD_DIR, OUTPUT_DIR]:
        if not folder.exists():
            continue

        for path in folder.iterdir():
            if path.is_file():
                try:
                    path.unlink()
                    deleted_count += 1
                except OSError:
                    pass

    return deleted_count


ensure_sample_files()


# =========================
# 文件安全工具函数
# =========================
def is_safe_filename(filename: str):
    """
    防止用户通过 ../ 之类的路径访问项目外部文件。
    只允许纯文件名，不允许路径分隔符。
    """
    if not filename:
        return False

    filename = str(filename)

    if "/" in filename or "\\" in filename:
        return False

    if filename in [".", ".."]:
        return False

    return True


def resolve_safe_file(folder: Path, filename: str):
    """
    将文件名安全地解析到指定目录下。
    如果文件不存在或路径非法，返回 None。
    """
    if not is_safe_filename(filename):
        return None

    try:
        base = folder.resolve()
        path = (folder / filename).resolve()

        if base not in path.parents and path != base:
            return None

        if not path.exists() or not path.is_file():
            return None

        return path

    except OSError:
        return None


# =========================
# 基础工具函数
# =========================
def safe_filename_part(text: str):
    text = str(text).strip()
    text = re.sub(r"[^\w\u4e00-\u9fff]+", "_", text)
    text = text.strip("_")
    return text or "column"


def auto_title(title: str, x_label: str, y_label: str):
    title = (title or "").strip()

    if title == "" or title.lower() == "x-y curve":
        return f"{y_label} vs {x_label}"

    return title


def read_csv_smart(file_path: Path, **kwargs):
    encodings = ["utf-8-sig", "utf-8", "gbk"]
    last_error = None

    for enc in encodings:
        try:
            return pd.read_csv(file_path, encoding=enc, **kwargs)
        except UnicodeDecodeError as exc:
            last_error = exc

    raise last_error


def read_raw_preview(file_path: Path, nrows: int = 15):
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        df = read_csv_smart(file_path, header=None, nrows=nrows)
    elif suffix == ".xlsx":
        df = pd.read_excel(file_path, engine="openpyxl", header=None, nrows=nrows)
    elif suffix == ".xls":
        df = pd.read_excel(file_path, engine="xlrd", header=None, nrows=nrows)
    else:
        raise ValueError("只支持 CSV、XLS、XLSX 文件。")

    df = df.fillna("")

    preview = []

    for idx, row in df.iterrows():
        preview.append({
            "row_number": idx + 1,
            "values": [str(v) for v in row.tolist()],
        })

    return preview


def clean_dataframe(df: pd.DataFrame):
    df = df.copy()

    cleaned_columns = []

    for idx, col in enumerate(df.columns):
        name = str(col).strip()

        if name == "" or name.lower().startswith("unnamed"):
            name = f"Column_{idx + 1}"

        cleaned_columns.append(name)

    df.columns = cleaned_columns
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")

    return df


def read_data_file(
    file_path: Path,
    header_row: int = 1,
    data_start_row: int | None = None,
    data_end_row: int | None = None,
):
    suffix = file_path.suffix.lower()
    header_index = header_row - 1

    if header_index < 0:
        raise ValueError("表头行必须大于等于第 1 行。")

    if data_start_row is None:
        data_start_row = header_row + 1

    if data_start_row <= header_row:
        raise ValueError("数据起始行必须在表头行之后。")

    if data_end_row is not None and data_end_row < data_start_row:
        raise ValueError("数据结束行不能早于数据起始行。")

    if suffix == ".csv":
        df = read_csv_smart(file_path, header=header_index)
    elif suffix == ".xlsx":
        df = pd.read_excel(file_path, engine="openpyxl", header=header_index)
    elif suffix == ".xls":
        df = pd.read_excel(file_path, engine="xlrd", header=header_index)
    else:
        raise ValueError("只支持 CSV、XLS、XLSX 文件。")

    first_data_row_number = header_row + 1
    start_offset = data_start_row - first_data_row_number

    if start_offset < 0:
        start_offset = 0

    if data_end_row is not None:
        end_offset = data_end_row - first_data_row_number + 1
        df = df.iloc[start_offset:end_offset]
    else:
        df = df.iloc[start_offset:]

    return clean_dataframe(df)


def get_numeric_columns(df: pd.DataFrame):
    numeric_columns = []

    for col in df.columns:
        converted = pd.to_numeric(df[col], errors="coerce")

        if converted.notna().sum() > 0:
            numeric_columns.append(col)

    return numeric_columns


def parse_positive_int(value, default=None):
    if value is None or str(value).strip() == "":
        return default

    try:
        num = int(value)
    except ValueError:
        raise ValueError(f"请输入有效的行号：{value}")

    if num <= 0:
        raise ValueError("行号必须是正整数。")

    return num


def parse_constant_k(value):
    if value is None or str(value).strip() == "":
        raise ValueError("当前计算类型需要填写常数 k。")

    try:
        return float(value)
    except ValueError:
        raise ValueError(f"常数 k 必须是数字：{value}")


def normalize_metric_mode(metric_mode: str):
    if metric_mode not in METRIC_MODES:
        return "basic"

    return metric_mode


def normalize_selected_metrics(metric_mode: str, selected_metrics: list[str] | None):
    metric_mode = normalize_metric_mode(metric_mode)

    if metric_mode == "basic":
        return BASIC_METRICS.copy()

    selected_metrics = selected_metrics or []
    cleaned = []

    for metric in selected_metrics:
        if metric in AVAILABLE_METRICS and metric not in cleaned:
            cleaned.append(metric)

    if not cleaned:
        cleaned = BASIC_METRICS.copy()

    return cleaned


def load_file_state(
    saved_filename: str,
    header_row: int,
    data_start_row: int,
    data_end_row: int | None,
):
    if not saved_filename:
        raise ValueError("文件状态丢失，请重新上传文件。")

    upload_path = resolve_safe_file(UPLOAD_DIR, saved_filename)

    if upload_path is None:
        raise ValueError("上传文件不存在或文件名非法，请重新上传。")

    preview = read_raw_preview(upload_path)

    df = read_data_file(
        upload_path,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
    )

    columns = [str(c) for c in df.columns]
    numeric_columns = get_numeric_columns(df)

    if len(numeric_columns) < 2:
        raise ValueError(
            f"至少需要两列数值数据才能绘图。当前可用数值列：{numeric_columns}"
        )

    return preview, columns, numeric_columns


# =========================
# 计算列函数：模板版
# =========================
def get_numeric_series(df: pd.DataFrame, col: str, label: str):
    if not col:
        raise ValueError(f"请选择{label}。")

    if col not in df.columns:
        raise ValueError(f"找不到{label}：{col}")

    series = pd.to_numeric(df[col], errors="coerce")

    if series.notna().sum() == 0:
        raise ValueError(f"{label}没有可用数值：{col}")

    return series


def calculate_template_column(
    df: pd.DataFrame,
    new_col_name: str,
    calc_template: str,
    first_col: str,
    second_col: str | None = None,
    constant_k: str | None = None,
):
    new_col_name = str(new_col_name or "").strip()

    if not new_col_name:
        raise ValueError("请输入新列名。")

    if new_col_name in df.columns:
        raise ValueError(f"新列名已经存在：{new_col_name}。请换一个名字。")

    if calc_template not in CALC_TEMPLATES:
        raise ValueError("请选择正确的计算类型。")

    a = get_numeric_series(df, first_col, "第一列 A")

    if calc_template in ["multiply", "divide", "add", "subtract"]:
        b = get_numeric_series(df, second_col, "第二列 B")

        if calc_template == "multiply":
            result = a * b
        elif calc_template == "divide":
            result = a / b
        elif calc_template == "add":
            result = a + b
        elif calc_template == "subtract":
            result = a - b
        else:
            raise ValueError("暂不支持该计算类型。")

    elif calc_template == "square":
        result = a ** 2

    elif calc_template == "add_const":
        k = parse_constant_k(constant_k)
        result = a + k

    elif calc_template == "subtract_const":
        k = parse_constant_k(constant_k)
        result = a - k

    elif calc_template == "multiply_const":
        k = parse_constant_k(constant_k)
        result = a * k

    elif calc_template == "divide_const":
        k = parse_constant_k(constant_k)

        if k == 0:
            raise ValueError("常数 k 不能为 0，否则会出现除以 0。")

        result = a / k

    elif calc_template == "sqrt":
        if (a < 0).any():
            raise ValueError("sqrt(A) 要求 A 不能包含负数。")

        result = np.sqrt(a)

    elif calc_template == "log10":
        if (a <= 0).any():
            raise ValueError("log10(A) 要求 A 必须全部大于 0。")

        result = np.log10(a)

    elif calc_template == "ln":
        if (a <= 0).any():
            raise ValueError("ln(A) 要求 A 必须全部大于 0。")

        result = np.log(a)

    elif calc_template == "abs":
        result = np.abs(a)

    else:
        raise ValueError("暂不支持该计算类型。")

    result = pd.Series(result)
    numeric_result = pd.to_numeric(result, errors="coerce")

    if numeric_result.replace([np.inf, -np.inf], np.nan).notna().sum() == 0:
        raise ValueError("计算结果没有可用数值，请检查所选列或常数。")

    if np.isinf(numeric_result).any():
        raise ValueError("计算结果出现无穷大，可能存在除以 0、log 非法输入或数值过大。")

    df = df.copy()
    df[new_col_name] = result

    return df, new_col_name


def save_processed_dataframe(df: pd.DataFrame, base_name: str = "processed"):
    job_id = uuid.uuid4().hex[:8]
    safe_base = safe_filename_part(base_name)

    filename = f"{safe_base}_{job_id}.csv"
    path = UPLOAD_DIR / filename

    df.to_csv(path, index=False, encoding="utf-8-sig")

    return filename


# =========================
# 拟合报告
# =========================
def save_fit_report(stats: dict, title: str = ""):
    job_id = uuid.uuid4().hex[:8]

    x_name = safe_filename_part(stats.get("x_col", "x"))
    y_name = safe_filename_part(stats.get("y_col", "y"))

    filename = f"{y_name}_vs_{x_name}_fit_report_{job_id}.txt"
    path = OUTPUT_DIR / filename

    lines = []

    lines.append("LabPlot Lite 拟合报告")
    lines.append("=" * 32)
    lines.append("")

    lines.append("一、图表信息")
    lines.append(f"图表标题：{title or '未填写'}")
    lines.append(f"X 数据列：{stats.get('x_col', '')}")
    lines.append(f"Y 数据列：{stats.get('y_col', '')}")
    lines.append(f"X 轴显示名称：{stats.get('x_label', '')}")
    lines.append(f"Y 轴显示名称：{stats.get('y_label', '')}")
    lines.append(f"图表类型：{stats.get('chart_type', '')}")
    lines.append(f"数据点数：{stats.get('points', '')}")
    lines.append("")

    lines.append("二、数据范围")
    lines.append(f"表头行：第 {stats.get('header_row', '')} 行")
    lines.append(f"数据起始行：第 {stats.get('data_start_row', '')} 行")
    lines.append(f"数据结束行：{stats.get('data_end_row', '')}")
    lines.append("")

    lines.append("三、基础统计")
    lines.append(f"Y 最大值：{stats.get('max_value', '')}")
    lines.append(f"Y 最小值：{stats.get('min_value', '')}")
    lines.append(f"Y 平均值：{stats.get('avg_value', '')}")
    lines.append(f"峰值对应 X：{stats.get('peak_x', '')}")
    lines.append("")

    lines.append("四、拟合结果")
    lines.append(f"拟合方式：{stats.get('fit_type_label', '不拟合')}")

    if stats.get("has_fit"):
        lines.append(f"拟合方程：{stats.get('equation', '')}")

        if stats.get("fit_type") == "linear":
            lines.append(f"斜率 a：{stats.get('fit_a', '')}")
            lines.append(f"截距 b：{stats.get('fit_b', '')}")

        elif stats.get("fit_type") == "quadratic":
            lines.append(f"二次项系数 a：{stats.get('fit_a', '')}")
            lines.append(f"一次项系数 b：{stats.get('fit_b', '')}")
            lines.append(f"常数项 c：{stats.get('fit_c', '')}")

        lines.append("")
        lines.append("五、拟合指标")

        selected_metrics = stats.get("selected_metrics", [])
        metric_values = stats.get("metric_values", {})

        if selected_metrics:
            for metric_key in selected_metrics:
                label = AVAILABLE_METRICS.get(metric_key, metric_key)
                value = metric_values.get(metric_key, "")
                lines.append(f"{label}：{value}")
        else:
            lines.append("未选择拟合指标。")

    else:
        lines.append("本次未进行拟合。")
        lines.append("")
        lines.append("五、拟合指标")
        lines.append("未进行拟合，因此没有拟合误差指标。")

    lines.append("")
    lines.append("六、提醒")
    lines.append("R² 只能反映拟合程度，不能单独证明物理模型正确。")
    lines.append("RMSE、MAE、最大绝对误差可以辅助判断拟合误差大小。")
    lines.append("请结合实验原理判断一次拟合或二次拟合是否合理。")

    path.write_text("\n".join(lines), encoding="utf-8")

    return filename


# =========================
# 拟合函数与误差指标
# =========================
def calc_r_squared(y_true: pd.Series, y_pred):
    ss_res = np.sum((y_true - y_pred) ** 2)
    ss_tot = np.sum((y_true - y_true.mean()) ** 2)

    if ss_tot == 0:
        return 1.0 if ss_res == 0 else 0.0

    return 1 - ss_res / ss_tot


def calc_all_fit_metrics(y_true: pd.Series, y_pred):
    residual = y_true - y_pred
    abs_residual = np.abs(residual)

    mse = np.mean(residual ** 2)
    rmse = np.sqrt(mse)
    mae = np.mean(abs_residual)
    max_abs_error = np.max(abs_residual)
    residual_mean = np.mean(residual)
    residual_std = np.std(residual, ddof=1) if len(residual) >= 2 else 0.0
    r2 = calc_r_squared(y_true, y_pred)

    return {
        "r2": r2,
        "mse": mse,
        "rmse": rmse,
        "mae": mae,
        "max_abs_error": max_abs_error,
        "residual_mean": residual_mean,
        "residual_std": residual_std,
    }


def format_metric_value(metric_key: str, value):
    if value is None:
        return ""

    return f"{value:.6f}"


def format_linear_equation(a, b):
    if b >= 0:
        return f"y = {a:.4f}x + {b:.4f}"

    return f"y = {a:.4f}x - {abs(b):.4f}"


def format_quadratic_equation(a, b, c):
    equation = f"y = {a:.4f}x²"

    if b >= 0:
        equation += f" + {b:.4f}x"
    else:
        equation += f" - {abs(b):.4f}x"

    if c >= 0:
        equation += f" + {c:.4f}"
    else:
        equation += f" - {abs(c):.4f}"

    return equation


def linear_fit(x: pd.Series, y: pd.Series):
    if len(x) < 2:
        raise ValueError("一次线性拟合至少需要 2 个有效数据点。")

    if x.nunique() < 2:
        raise ValueError("X 轴数据至少需要两个不同的数值，才能进行一次线性拟合。")

    a, b = np.polyfit(x, y, 1)
    y_pred = a * x + b
    all_metrics = calc_all_fit_metrics(y, y_pred)

    x_fit = np.linspace(x.min(), x.max(), 200)
    y_fit = a * x_fit + b

    return {
        "type": "linear",
        "a": a,
        "b": b,
        "c": None,
        "y_pred": y_pred,
        "all_metrics": all_metrics,
        "x_fit": x_fit,
        "y_fit": y_fit,
        "equation": format_linear_equation(a, b),
    }


def quadratic_fit(x: pd.Series, y: pd.Series):
    if len(x) < 3:
        raise ValueError("二次拟合至少需要 3 个有效数据点。")

    if x.nunique() < 3:
        raise ValueError("X 轴数据至少需要三个不同的数值，才能进行二次拟合。")

    a, b, c = np.polyfit(x, y, 2)
    y_pred = a * x ** 2 + b * x + c
    all_metrics = calc_all_fit_metrics(y, y_pred)

    x_fit = np.linspace(x.min(), x.max(), 300)
    y_fit = a * x_fit ** 2 + b * x_fit + c

    return {
        "type": "quadratic",
        "a": a,
        "b": b,
        "c": c,
        "y_pred": y_pred,
        "all_metrics": all_metrics,
        "x_fit": x_fit,
        "y_fit": y_fit,
        "equation": format_quadratic_equation(a, b, c),
    }


# =========================
# 绘图函数
# =========================
def make_xy_plot(
    data_path: Path,
    x_col: str,
    y_col: str,
    title: str = "",
    x_label: str = "",
    y_label: str = "",
    header_row: int = 1,
    data_start_row: int | None = None,
    data_end_row: int | None = None,
    chart_type: str = "line_marker",
    fit_type: str = "none",
    metric_mode: str = "basic",
    selected_metrics: list[str] | None = None,
):
    metric_mode = normalize_metric_mode(metric_mode)
    selected_metrics = normalize_selected_metrics(metric_mode, selected_metrics)

    df = read_data_file(
        data_path,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
    )

    if x_col not in df.columns:
        raise ValueError(f"找不到 X 轴列：{x_col}")

    if y_col not in df.columns:
        raise ValueError(f"找不到 Y 轴列：{y_col}")

    if chart_type not in CHART_TYPES:
        raise ValueError(f"不支持的图表类型：{chart_type}")

    if fit_type not in FIT_TYPES:
        raise ValueError(f"不支持的拟合方式：{fit_type}")

    x = pd.to_numeric(df[x_col], errors="coerce")
    y = pd.to_numeric(df[y_col], errors="coerce")

    clean = pd.DataFrame({
        x_col: x,
        y_col: y,
    }).dropna()

    if clean.empty:
        raise ValueError("所选 X/Y 列中没有可用的数值数据，请检查表格内容。")

    x = clean[x_col]
    y = clean[y_col]

    display_x_label = (x_label or "").strip() or str(x_col)
    display_y_label = (y_label or "").strip() or str(y_col)

    max_value = y.max()
    min_value = y.min()
    avg_value = y.mean()
    max_x = clean.loc[y.idxmax(), x_col]

    fit_result = None

    if fit_type == "linear":
        fit_result = linear_fit(x, y)
    elif fit_type == "quadratic":
        fit_result = quadratic_fit(x, y)

    job_id = uuid.uuid4().hex[:8]
    x_name = safe_filename_part(x_col)
    y_name = safe_filename_part(y_col)
    prefix = f"{y_name}_vs_{x_name}"

    png_path = OUTPUT_DIR / f"{prefix}_{job_id}.png"
    origin_csv_path = OUTPUT_DIR / f"{prefix}_origin_{job_id}.csv"
    full_csv_path = OUTPUT_DIR / f"{prefix}_full_data_{job_id}.csv"

    origin_df = clean.copy()

    metric_values = {}
    metric_display = []

    if fit_result is not None:
        if fit_type == "linear":
            fit_y_col = "linear_fit_y"
        elif fit_type == "quadratic":
            fit_y_col = "quadratic_fit_y"
        else:
            fit_y_col = "fit_y"

        origin_df[fit_y_col] = fit_result["y_pred"]
        origin_df["residual"] = origin_df[y_col] - origin_df[fit_y_col]
        origin_df["abs_residual"] = np.abs(origin_df["residual"])

        all_metrics = fit_result["all_metrics"]

        for metric_key in selected_metrics:
            raw_value = all_metrics.get(metric_key)
            formatted_value = format_metric_value(metric_key, raw_value)
            metric_values[metric_key] = formatted_value
            metric_display.append({
                "key": metric_key,
                "label": AVAILABLE_METRICS.get(metric_key, metric_key),
                "value": formatted_value,
            })

    origin_df.to_csv(origin_csv_path, index=False, encoding="utf-8-sig")
    df.to_csv(full_csv_path, index=False, encoding="utf-8-sig")

    plot_title = auto_title(title, display_x_label, display_y_label)

    fig, ax = plt.subplots(figsize=(6.5, 4.2), dpi=150)

    if fit_result is not None:
        ax.scatter(
            x,
            y,
            s=30,
            label="Data",
        )

        fit_label = "Linear Fit" if fit_type == "linear" else "Quadratic Fit"

        ax.plot(
            fit_result["x_fit"],
            fit_result["y_fit"],
            linewidth=1.8,
            label=fit_label,
        )

        r2_text = metric_values.get(
            "r2",
            format_metric_value("r2", fit_result["all_metrics"].get("r2")),
        )

        fit_text = f"{fit_result['equation']}\nR² = {r2_text}"

        ax.text(
            0.05,
            0.95,
            fit_text,
            transform=ax.transAxes,
            fontsize=9,
            verticalalignment="top",
            bbox=dict(
                boxstyle="round,pad=0.3",
                facecolor="white",
                alpha=0.75,
                edgecolor="#cccccc",
            ),
        )

    else:
        if chart_type == "line":
            ax.plot(
                x,
                y,
                linewidth=1.8,
                label=str(y_col),
            )

        elif chart_type == "scatter":
            ax.scatter(
                x,
                y,
                s=28,
                label=str(y_col),
            )

        elif chart_type == "line_marker":
            ax.plot(
                x,
                y,
                marker="o",
                linewidth=1.8,
                markersize=4.5,
                label=str(y_col),
            )

    ax.set_title(plot_title, fontsize=13, pad=10)
    ax.set_xlabel(display_x_label, fontsize=11)
    ax.set_ylabel(display_y_label, fontsize=11)
    ax.tick_params(axis="both", labelsize=9)
    ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.35)
    ax.legend(fontsize=9, frameon=True)

    if len(clean) >= 2:
        x_margin = (x.max() - x.min()) * 0.05 if x.max() != x.min() else 1
        y_margin = (y.max() - y.min()) * 0.12 if y.max() != y.min() else 1

        ax.set_xlim(x.min() - x_margin, x.max() + x_margin)
        ax.set_ylim(y.min() - y_margin, y.max() + y_margin)

    fig.tight_layout()
    fig.savefig(png_path, dpi=300)
    plt.close(fig)

    stats = {
        "x_col": str(x_col),
        "y_col": str(y_col),
        "x_label": display_x_label,
        "y_label": display_y_label,
        "max_value": f"{max_value:.2f}",
        "min_value": f"{min_value:.2f}",
        "avg_value": f"{avg_value:.2f}",
        "peak_x": f"{max_x}",
        "points": str(len(clean)),
        "header_row": str(header_row),
        "data_start_row": str(data_start_row),
        "data_end_row": str(data_end_row) if data_end_row else "未限制",
        "chart_type": CHART_TYPES.get(chart_type, chart_type),
        "fit_type": fit_type,
        "fit_type_label": FIT_TYPES.get(fit_type, fit_type),
        "has_fit": fit_result is not None,
        "equation": None,
        "fit_a": None,
        "fit_b": None,
        "fit_c": None,
        "metric_mode": metric_mode,
        "metric_mode_label": METRIC_MODES.get(metric_mode, metric_mode),
        "selected_metrics": selected_metrics,
        "metric_values": metric_values,
        "metric_display": metric_display,
    }

    if fit_result is not None:
        stats["chart_type"] = FIT_TYPES.get(fit_type, fit_type)
        stats["equation"] = fit_result["equation"]
        stats["fit_a"] = f"{fit_result['a']:.6f}"
        stats["fit_b"] = f"{fit_result['b']:.6f}"

        if fit_type == "quadratic":
            stats["fit_c"] = f"{fit_result['c']:.6f}"

    fit_report_name = save_fit_report(stats, plot_title)

    return png_path.name, origin_csv_path.name, full_csv_path.name, fit_report_name, stats


# =========================
# Flask 路由
# =========================
@app.route("/", methods=["GET", "POST"])
def index():
    result = None
    error = None
    success = None

    sample_files = SAMPLE_FILES

    preview = None
    columns = None
    numeric_columns = None

    upload_name = None
    saved_filename = None

    selected_x = None
    selected_y = None

    title = "X-Y Curve"
    x_label = ""
    y_label = ""

    chart_type = "line_marker"
    fit_type = "none"

    metric_mode = "basic"
    selected_metrics = BASIC_METRICS.copy()

    header_row = 1
    data_start_row = 2
    data_end_row = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "cleanup":
            deleted_count = cleanup_temp_files()
            success = f"已清理临时文件：{deleted_count} 个。"

        elif action == "upload":
            file = request.files.get("csv_file")

            if not file or file.filename == "":
                error = "请先上传 CSV / Excel 文件。"

            elif Path(file.filename).suffix.lower() not in ALLOWED_EXTENSIONS:
                error = "目前只支持 CSV、XLS、XLSX 文件。"

            else:
                try:
                    original_suffix = Path(file.filename).suffix.lower()
                    safe_name = f"{uuid.uuid4().hex}{original_suffix}"
                    upload_path = UPLOAD_DIR / safe_name
                    file.save(upload_path)

                    saved_filename = safe_name
                    upload_name = file.filename

                    preview, columns, numeric_columns = load_file_state(
                        saved_filename,
                        header_row,
                        data_start_row,
                        data_end_row,
                    )

                    selected_x = numeric_columns[0]
                    selected_y = numeric_columns[1]

                except Exception as exc:
                    error = str(exc)

        elif action == "reload":
            try:
                saved_filename = request.form.get("saved_filename")
                upload_name = request.form.get("upload_name")

                header_row = parse_positive_int(request.form.get("header_row"), 1)
                data_start_row = parse_positive_int(
                    request.form.get("data_start_row"),
                    header_row + 1,
                )
                data_end_row = parse_positive_int(request.form.get("data_end_row"), None)

                preview, columns, numeric_columns = load_file_state(
                    saved_filename,
                    header_row,
                    data_start_row,
                    data_end_row,
                )

                selected_x = numeric_columns[0]
                selected_y = numeric_columns[1]

            except Exception as exc:
                error = str(exc)

        elif action == "calc":
            try:
                saved_filename = request.form.get("saved_filename")
                upload_name = request.form.get("upload_name")

                header_row = parse_positive_int(request.form.get("header_row"), 1)
                data_start_row = parse_positive_int(
                    request.form.get("data_start_row"),
                    header_row + 1,
                )
                data_end_row = parse_positive_int(request.form.get("data_end_row"), None)

                new_col_name = request.form.get("new_col_name")
                calc_template = request.form.get("calc_template")
                first_col = request.form.get("first_col")
                second_col = request.form.get("second_col")
                constant_k = request.form.get("constant_k")

                if not saved_filename:
                    raise ValueError("文件状态丢失，请重新上传文件。")

                upload_path = resolve_safe_file(UPLOAD_DIR, saved_filename)

                if upload_path is None:
                    raise ValueError("上传文件不存在或文件名非法，请重新上传。")

                df = read_data_file(
                    upload_path,
                    header_row=header_row,
                    data_start_row=data_start_row,
                    data_end_row=data_end_row,
                )

                df, created_col = calculate_template_column(
                    df,
                    new_col_name=new_col_name,
                    calc_template=calc_template,
                    first_col=first_col,
                    second_col=second_col,
                    constant_k=constant_k,
                )

                saved_filename = save_processed_dataframe(df, created_col)
                upload_name = f"{upload_name} + 新列 {created_col}"

                header_row = 1
                data_start_row = 2
                data_end_row = None

                preview, columns, numeric_columns = load_file_state(
                    saved_filename,
                    header_row,
                    data_start_row,
                    data_end_row,
                )

                selected_x = numeric_columns[0]
                selected_y = created_col if created_col in numeric_columns else numeric_columns[1]

                success = f"已生成新列：{created_col}"

            except Exception as exc:
                error = str(exc)

        elif action == "plot":
            try:
                saved_filename = request.form.get("saved_filename")
                upload_name = request.form.get("upload_name")

                header_row = parse_positive_int(request.form.get("header_row"), 1)
                data_start_row = parse_positive_int(
                    request.form.get("data_start_row"),
                    header_row + 1,
                )
                data_end_row = parse_positive_int(request.form.get("data_end_row"), None)

                selected_x = request.form.get("x_col")
                selected_y = request.form.get("y_col")

                title = request.form.get("title", "X-Y Curve").strip()
                x_label = request.form.get("x_label", "").strip()
                y_label = request.form.get("y_label", "").strip()

                chart_type = request.form.get("chart_type", "line_marker")
                fit_type = request.form.get("fit_type", "none")

                metric_mode = normalize_metric_mode(request.form.get("metric_mode", "basic"))
                selected_metrics = normalize_selected_metrics(
                    metric_mode,
                    request.form.getlist("selected_metrics"),
                )

                if not saved_filename:
                    raise ValueError("文件状态丢失，请重新上传文件。")

                if not selected_x or not selected_y:
                    raise ValueError("请选择 X 轴和 Y 轴。")

                if selected_x == selected_y:
                    raise ValueError("X 轴和 Y 轴不能选择同一列。")

                if chart_type not in CHART_TYPES:
                    raise ValueError("请选择正确的图表类型。")

                if fit_type not in FIT_TYPES:
                    raise ValueError("请选择正确的拟合方式。")

                preview, columns, numeric_columns = load_file_state(
                    saved_filename,
                    header_row,
                    data_start_row,
                    data_end_row,
                )

                upload_path = resolve_safe_file(UPLOAD_DIR, saved_filename)

                if upload_path is None:
                    raise ValueError("上传文件不存在或文件名非法，请重新上传。")

                png_name, origin_csv_name, full_csv_name, fit_report_name, stats = make_xy_plot(
                    upload_path,
                    selected_x,
                    selected_y,
                    title,
                    x_label=x_label,
                    y_label=y_label,
                    header_row=header_row,
                    data_start_row=data_start_row,
                    data_end_row=data_end_row,
                    chart_type=chart_type,
                    fit_type=fit_type,
                    metric_mode=metric_mode,
                    selected_metrics=selected_metrics,
                )

                result = {
                    "png_name": png_name,
                    "origin_csv_name": origin_csv_name,
                    "full_csv_name": full_csv_name,
                    "fit_report_name": fit_report_name,
                    "stats": stats,
                }

            except Exception as exc:
                error = str(exc)

        else:
            error = "未知操作，请重新上传文件。"

    return render_template(
        "index.html",
        result=result,
        error=error,
        success=success,
        sample_files=sample_files,
        preview=preview,
        columns=columns,
        numeric_columns=numeric_columns,
        upload_name=upload_name,
        saved_filename=saved_filename,
        selected_x=selected_x,
        selected_y=selected_y,
        title=title,
        x_label=x_label,
        y_label=y_label,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        chart_type=chart_type,
        chart_types=CHART_TYPES,
        fit_type=fit_type,
        fit_types=FIT_TYPES,
        metric_mode=metric_mode,
        metric_modes=METRIC_MODES,
        selected_metrics=selected_metrics,
        available_metrics=AVAILABLE_METRICS,
        basic_metrics=BASIC_METRICS,
        calc_templates=CALC_TEMPLATES,
    )


@app.route("/sample/<filename>")
def download_sample(filename):
    if filename not in SAMPLE_FILES:
        return "示例文件不存在", 404

    path = resolve_safe_file(SAMPLE_DIR, filename)

    if path is None:
        ensure_sample_files()
        path = resolve_safe_file(SAMPLE_DIR, filename)

    if path is None:
        return "示例文件生成失败", 500

    return send_file(path, as_attachment=True)


@app.route("/download/<filename>")
def download(filename):
    path = resolve_safe_file(OUTPUT_DIR, filename)

    if path is None:
        return "文件不存在", 404

    return send_file(path, as_attachment=True)


@app.route("/view/<filename>")
def view_file(filename):
    path = resolve_safe_file(OUTPUT_DIR, filename)

    if path is None:
        return "文件不存在", 404

    return send_file(path)


if __name__ == "__main__":
    print("=" * 48)
    print("LabPlot Lite 已启动")
    print("本地访问：http://127.0.0.1:5000")
    print("当前模式：内测安全模式，debug=False")
    print("=" * 48)

    app.run(
        host="127.0.0.1",
        port=5000,
        debug=False,
    )